import pandas as pd
from datetime import datetime, timedelta
from typing import Optional
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

class ExcelExporter:
    
    @staticmethod
    def export_posture_data_to_excel(data: pd.DataFrame, filename: str = None, 
                                   start_date: Optional[datetime] = None, 
                                   end_date: Optional[datetime] = None) -> str:
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"posture_data_{timestamp}.xlsx"
        
        if start_date or end_date:
            data = data.copy()
            data['timestamp'] = pd.to_datetime(data['timestamp'])
            
            if start_date:
                data = data[data['timestamp'] >= start_date]
            if end_date:
                data = data[data['timestamp'] <= end_date]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Posture Data"
        
        headers = []
        if 'id' in data.columns:
            headers.append('ID')
        if 'timestamp' in data.columns:
            headers.append('Timestamp')
        if 'shoulder_angle' in data.columns:
            headers.append('Shoulder Angle')
        elif 'angulo_ombro' in data.columns:
            headers.append('Shoulder Angle')
        if 'neck_angle' in data.columns:
            headers.append('Neck Angle')
        elif 'angulo_pescoco' in data.columns:
            headers.append('Neck Angle')
        if 'camera_type' in data.columns:
            headers.append('Camera')
        elif 'camera' in data.columns:
            headers.append('Camera')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row_idx, (_, row_data) in enumerate(data.iterrows(), 2):
            col_idx = 1
            
            if 'id' in data.columns:
                ws.cell(row=row_idx, column=col_idx, value=row_data['id'])
                col_idx += 1
            
            if 'timestamp' in data.columns:
                timestamp = pd.to_datetime(row_data['timestamp'])
                ws.cell(row=row_idx, column=col_idx, value=timestamp.strftime("%Y-%m-%d %H:%M:%S"))
                col_idx += 1
            
            if 'shoulder_angle' in data.columns:
                ws.cell(row=row_idx, column=col_idx, value=row_data['shoulder_angle'])
                col_idx += 1
            elif 'angulo_ombro' in data.columns:
                ws.cell(row=row_idx, column=col_idx, value=row_data['angulo_ombro'])
                col_idx += 1
            
            if 'neck_angle' in data.columns:
                ws.cell(row=row_idx, column=col_idx, value=row_data['neck_angle'])
                col_idx += 1
            elif 'angulo_pescoco' in data.columns:
                ws.cell(row=row_idx, column=col_idx, value=row_data['angulo_pescoco'])
                col_idx += 1
            
            if 'camera_type' in data.columns:
                camera = row_data['camera_type'].title()
                ws.cell(row=row_idx, column=col_idx, value=camera)
                col_idx += 1
            elif 'camera' in data.columns:
                camera = row_data['camera'].title()
                ws.cell(row=row_idx, column=col_idx, value=camera)
                col_idx += 1
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        summary_ws = wb.create_sheet("Summary")
        summary_ws.title = "Summary"
        
        summary_ws['A1'] = "Posture Data Summary"
        summary_ws['A1'].font = Font(bold=True, size=14)
        
        summary_ws['A3'] = "Total Records:"
        summary_ws['B3'] = len(data)
        
        if 'timestamp' in data.columns:
            data['timestamp'] = pd.to_datetime(data['timestamp'])
            summary_ws['A4'] = "Date Range:"
            summary_ws['B4'] = f"{data['timestamp'].min().strftime('%Y-%m-%d')} to {data['timestamp'].max().strftime('%Y-%m-%d')}"
        
        if 'camera_type' in data.columns or 'camera' in data.columns:
            camera_col = 'camera_type' if 'camera_type' in data.columns else 'camera'
            camera_counts = data[camera_col].value_counts()
            summary_ws['A6'] = "Camera Distribution:"
            summary_ws['A6'].font = Font(bold=True)
            
            for idx, (camera, count) in enumerate(camera_counts.items(), 7):
                summary_ws[f'A{idx}'] = f"{camera.title()}:"
                summary_ws[f'B{idx}'] = count
        
        filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)
        
        return filepath
    
    @staticmethod
    def get_export_filename(start_date: Optional[datetime] = None, 
                          end_date: Optional[datetime] = None) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if start_date and end_date:
            start_str = start_date.strftime("%Y%m%d")
            end_str = end_date.strftime("%Y%m%d")
            return f"posture_data_{start_str}_to_{end_str}_{timestamp}.xlsx"
        elif start_date:
            start_str = start_date.strftime("%Y%m%d")
            return f"posture_data_from_{start_str}_{timestamp}.xlsx"
        elif end_date:
            end_str = end_date.strftime("%Y%m%d")
            return f"posture_data_until_{end_str}_{timestamp}.xlsx"
        else:
            return f"posture_data_all_{timestamp}.xlsx" 